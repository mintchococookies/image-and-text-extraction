import cv2
import numpy as np
import pytesseract as tess
import openpyxl

# Detect the type of image by checking if there are two large areas in the image
def getType(img):
    [nrow, ncol] = img.shape
    img_copy = np.pad(img[11:nrow-11,11:ncol-11], pad_width = 50, mode = 'constant', constant_values = 255)
    ret, bi_img = cv2.threshold(img_copy, 120, 1, cv2.THRESH_BINARY)
    erosion_sE = cv2.getStructuringElement(cv2.MORPH_RECT,(50,50))
    erosion_output = cv2.erode(bi_img, erosion_sE)
    contours, hierarchy = cv2.findContours(erosion_output, cv2.RETR_TREE, cv2.CHAIN_APPROX_NONE)
    
    #Check for areas that are greater than the threshold area
    threshold_area = 2000000
    count = 0     
    for cnt in contours:        
        area = cv2.contourArea(cnt)         
        if area > threshold_area: 
            count += 1      
            
    return count

# Get the coordinates of the area around all items in the image
def getContourBoundingBox(img):
    contours, hierarchy = cv2.findContours(img, cv2.RETR_TREE, cv2.CHAIN_APPROX_NONE)
    [min_x, min_y] = img.shape
    max_x = max_y = 0
    for contour in contours:
        x, y, w, h = cv2.boundingRect(contour)
        if x != 0 and y != 0 and w !=0 and h != 0:
            min_x, max_x = min(x, min_x), max(x + w, max_x)
            min_y, max_y = min(y, min_y), max(y + h, max_y)
    return min_x, min_y, max_x, max_y

# Return the cropped drawing and inverse binary image with only the table
def getComponents(img, imgtype):
    ret, inv_bi_img = cv2.threshold(img, 100, 1, cv2.THRESH_BINARY_INV)
    [nrow,ncol] = img.shape
    
    #First Method: Erosion & Dilation (for images 11-20)
    if imgtype == 1:
        
        #Erode until only the thickest parts remain, then dilate and erode to get a binary mask
        erosion_sE = cv2.getStructuringElement(cv2.MORPH_RECT,(9,9))
        erosion_output = cv2.erode(inv_bi_img, erosion_sE)
        dilation_sE = cv2.getStructuringElement(cv2.MORPH_RECT,(64,64))
        dilation_output = cv2.dilate(erosion_output, dilation_sE, anchor = (-1, -1), iterations = 9)
        mask = cv2.erode(dilation_output, erosion_sE, iterations = 40)
    
        #Add padding at the top of the region of interest to include the "VIEW" labels
        min_x, min_y, max_x, max_y = getContourBoundingBox(mask)
        mask[20:min_y+2, min_x:max_x] = 1
        table_mask = mask.copy()
        table_mask[max_y-80:,min_x:max_x] = 0
        
        #Bitwise AND the image with the created mask to extract the drawing and table
        drawing = cv2.bitwise_and(mask, inv_bi_img)
        table = cv2.bitwise_and((1-table_mask), inv_bi_img)

        #Crop the inverted drawing image based on the 4 corners
        drawing = cv2.bitwise_not(cv2.threshold(drawing, 0, 255, cv2.THRESH_OTSU)[1])
        min_y -= 250
        max_y += 100
        if min_y < 10: 
            min_y = 10
        cropped_drawing = drawing[min_y:max_y, min_x:max_x]
        
    #Second Method: Contours with Bounding Boxes (Images 1-10)
    elif imgtype >= 2:
        
        #Erode the image until only the thickest parts remain
        ret, inv_bi_img_copy = cv2.threshold(img[10:nrow-10, 10:ncol-10], 100, 1, cv2.THRESH_BINARY_INV)
        erosion_sE = cv2.getStructuringElement(cv2.MORPH_RECT, (10,10))
        erosion_output = cv2.erode(inv_bi_img_copy, erosion_sE, anchor = (-1, -1), iterations = 1)
        
        #Find the contours of drawing and get their bounding boxes
        min_x, min_y, max_x, max_y = getContourBoundingBox(erosion_output)
        min_x -= 130
        min_y -= 250
        max_x += 130
        max_y += 60
        if min_y < 10: 
            min_y = 10
        
        #Create a mask based on the padded bounding box and bitwise AND with the image to extract the table
        mask = inv_bi_img.copy()
        mask[min_y:max_y, min_x:max_x] = 0
        table = cv2.bitwise_and(mask, inv_bi_img)
        
        #Crop the inverted drawing image based on the 4 corners
        cropped_drawing = img[min_y:max_y, min_x:max_x]

    return cropped_drawing, table

# Extract text from the table image
def extractText(table, j):
    #Remove the vertical and horizontal lines from the table
    horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (55,1))
    remove_horizontal = cv2.morphologyEx(table, cv2.MORPH_OPEN, horizontal_kernel)
    vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1,55))
    remove_vertical = cv2.morphologyEx(table, cv2.MORPH_OPEN, vertical_kernel)
    table_wo_lines = table * (1-remove_horizontal) * (1-remove_vertical)
    
    #Dilate the images for better extraction
    dilation_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))   
    dil_table_wo_lines = cv2.dilate(table_wo_lines, dilation_kernel, iterations=1)
    dil_table_wo_lines = cv2.bitwise_not(cv2.threshold(dil_table_wo_lines, 0, 255, cv2.THRESH_OTSU)[1]) #table without lines
    
    #Find and extract the Amendments table first
    data = tess.image_to_data(dil_table_wo_lines, output_type=tess.Output.DICT)
    nrow, ncol = table.shape
    min_x = max_x = min_y = max_y = 0
    for i in range(len(data['text'])):   
        if data['text'][i] == "AMENDMENTS":
            left, top, width, height = data['left'][i], data['top'][i], data['width'][i], data['height'][i]  
            table[top:top + height, left:left + width] = 0
            #Check to the left of "AMENDMENTS" to find the table's left edge  
            for pixel in range(left, 0, -1): 
                if table[top, pixel] == 1:
                    min_x = pixel
                    break
            #Check to the right of "AMENDMENTS" to find the table's right edge  
            for pixel in range(left, ncol): 
                if table[top, pixel] == 1:
                    max_x = pixel
                    break
            #Check to the top of "AMENDMENTS" to find the table's top edge  
            for pixel in range(top, 0, -1):
                if table[pixel, left] == 1:
                    min_y = pixel
                    break
            #cCheck to the bottom of "AMENDMENTS" to find the end of the row
            for pixel in range(top, nrow): 
                if table[pixel, left] == 1:
                    temp_y = pixel
                    break
            #Allow a maximum of 3 rows to the table's bottom edge
            for pixel in range(temp_y+5, nrow): 
                if table[pixel, left] == 1:
                    max_y = pixel
                    max_y = max_y + ((max_y-temp_y) * 3)
                    break
    amend_table = table_wo_lines[min_y:max_y,min_x:max_x]
    
    #Dilate the amendments table separately for better accuracy
    dilation_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 2))   
    amend_table = cv2.dilate(amend_table, dilation_kernel,iterations=1)
    amend_table = cv2.bitwise_not(cv2.threshold(amend_table, 0, 255, cv2.THRESH_OTSU)[1])
    
    #Extract and sort the data from the Amendments table
    amend_data = tess.image_to_string(amend_table)
    amend_data = amend_data.split("\n") 
    amend_data.pop(0)
    
    while '' in amend_data: 
        amend_data.remove('')
    for k in range(0, len(amend_data)):
        if len(amend_data[k]) > 1: 
            amend_data[k] = amend_data[k].split(" ")
        else: amend_data.pop(k)
    
    #Join the multi-word data if the headings list is shorter than the data list
    if len(amend_data[0]) != len(amend_data[1]): 
        for ad in range(1, len(amend_data)):
            amend_data[ad][2] = ' '.join(amend_data[ad][2:-2])
            del amend_data[ad][3:-2]
    
    #Extract the other parts of the table
    dil_table_wo_lines[min_y:max_y, min_x:max_x] = 255
    data = tess.image_to_string(dil_table_wo_lines)
    data = data.split("\n")
    
    #Scan the extracted data list to separate the headings and data
    possible_headings = ["TITLE","DRAWING N","PROJECT NO","CAD NO","COMPANY","UNIT","PAGE",
                         "STATUS","DRAWN","CHECKED","APPROVED","CONTRACTOR","FONT","LANG"]
    headings_list = []
    data_list = []
    for d in range(len(data)):
        for h in range(len(possible_headings)):
            if possible_headings[h] in data[d]: 
                headings_list.append(str(data[d]))
                data[d] = ''
        if data[d] != ' ' and data[d] != '' and data[d] != "BOTTOM VIEW": 
            data_list.append(str(data[d]))
    
    #Split the heading and data lists into individual headings and data
    for hl in range(len(headings_list)):
        headings_list[hl] = headings_list[hl].split(': ')
    for dl in range(len(data_list)):
        data_list[dl] = data_list[dl].split()       
    
    #Combine multi-word data into one word
    for hl in range(len(headings_list)):
        for h in range(len(headings_list[hl])):
            if "PROJECT NO" in (headings_list[hl][h]) and data_list[hl][h].count('-') == 0:
                if data_list[hl][h] == "E":
                    data_list[hl][h] = '-'.join(data_list[hl][h:h+4])
                    del data_list[hl][h+1:h+4]
                else:
                    data_list[hl][h] = ' '.join(data_list[hl][h:h+3])
                    del data_list[hl][h+1:h+3]
            elif "CAD" in (headings_list[hl][h]) and len(data_list[hl]) > 3:
                data_list[hl][h] = ' '.join(data_list[hl][h:h+4])
                del data_list[hl][h+1:h+4]
            elif "DRAWING N" in (headings_list[hl][h]) and len(data_list[hl]) > 7:
                data_list[hl][h] = ' '.join(data_list[hl][h:h+8])
                del data_list[hl][h+1:h+8]
            elif "TITLE" in (headings_list[hl][h]) or "FONT" in (headings_list[hl][h]) or "CONTRACTOR" in (headings_list[hl][h]):
                if (len(headings_list[hl])-1) - h == 0:
                    data_list[hl][h] = ' '.join(data_list[hl][h:])
                    del data_list[hl][h+1:]
                else:
                    data_list[hl][h] = ' '.join(data_list[hl][h:(-((len(headings_list[hl])-1) - h))])
                    del data_list[hl][h+1:(-((len(headings_list[hl])-1) - h))]
    
    #Insert the data into the correct indexes in the results array
    results = []
    for ph in range(len(possible_headings)):
        results_length = len(results)
        for hl in range(len(headings_list)):
            for h in range(len(headings_list[hl])): 
                if possible_headings[ph] in (headings_list[hl][h]):                  
                    results.append(data_list[hl][h])          
        if len(results) == results_length:
            results.append("Unavailable") 
    results.append(amend_data)
    
    return results

# Create an Excel spreadsheet with all the data
def generateExcel(array):
    wb = openpyxl.Workbook()
    ws = wb.active
    
    #Insert the column headings into the spreadsheet
    excel_headings = ["NO.","TITLE","DRAWING NO","PROJECT NO","CAD NO","COMPANY","UNIT","PAGE","STATUS",
                      "DRAWN BY","CHECKED BY","APPROVED BY","CONTRACTOR","FONT","LANGUAGE","AMENDMENTS"]
    for y in range(0,len(excel_headings)):
        ws.cell(row = 1, column = y+1).value = excel_headings[y]
        ws.cell(row = 1, column = y+1).font = openpyxl.styles.Font(bold = True)
    
    #Insert the extracted data into the spreadsheet
    r = 2
    for x in range(0,len(excel_results)):
        for y in range(0,len(excel_results[x])):
            ws.cell(row = r, column = 1).value = x+1 
            if y == 14:
                for m in range(0,len(excel_results[x][14])):
                    for n in range(0,len(excel_results[x][14][m])):
                        ws.cell(row=r+m,column=y+2+n).value = excel_results[x][14][m][n]
                        ws.cell(row=r,column=y+2+n).font = openpyxl.styles.Font(bold = True)
            else:
                ws.cell(row = r,column = y+2).value = excel_results[x][y]
            ws.merge_cells(start_row = r,start_column = y+1, end_row = r-1 + len(excel_results[x][14]), end_column = y+1)
        r += len(excel_results[x][-1])
    
    #Adjust the width of the cells
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length * 1.4
    
    #Merge the columns for the Amendments heading
    ws.merge_cells(start_row = 1, start_column = 16, end_row = 1, end_column = 20)
    
    wb.save(filename="drawinginfo.xlsx")   


### ============ MAIN ============ ###

excel_results = []

files = ["01.png","02.png","03.png","04.png","05.png","06.png","07.png","08.png","09.png","10.png",
          "11.png","12.png","13.png","14.png","15.png","16.png","17.png","18.png","19.png","20.png"]

#Loop through all 20 images
for j in range(0, 20):
    img = cv2.imread(files[j], 0)
    print("Processing Image " + str(j+1))
    
    #Get the separated components and save the drawing as a PNG file
    cropped_drawing, table = getComponents(img, getType(img))
    cv2.imwrite("Drawing {}.png".format(j+1), cropped_drawing)
    
    #Extract the text from the table
    try:
        excel_results.append(extractText(table, j))
    except:
        excel_results.append("Error Processing Image")
        print("Error Processing Image")

generateExcel(excel_results)

print("Completed Processing " + str(j+1) + " Images!")